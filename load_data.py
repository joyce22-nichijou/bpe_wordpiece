"""
load_data.py
============
Loads train/test corpora (WikiText + Gutenberg) and two gold-standard
morphology resources (MorphoLex, Morpho Challenge GoldStd).

WikiText / Gutenberg are downloaded once (via `datasets`) and cached to
data/*.txt, so a fresh checkout (e.g. on the grader's machine) is fully
reproducible without committing multi-hundred-MB corpus files to git.
Both are sliced into small/medium/large size tiers (see CORPUS_SIZES) to
compare training-corpus scale.

MorphoLex   -> canonical-form morphemes + prefix/root/suffix roles,
               used by affix_detection / segmentation_consistency.
GoldStd     -> surface-form morphemes, 100% character-aligned,
               used by boundary_prf / morpheme_recovery / over_under_segmentation.
"""

from __future__ import annotations
import re
import pickle
from pathlib import Path
from collections import defaultdict

from tokenizer_interface import load_corpus

# Parsed MorphoLex/GoldStd databases are pickled here after the first parse
# (the MorphoLex .xlsx parse in particular is slow); reused while the source
# file is unchanged, so repeated run_eval.py sessions don't re-parse them.
_CACHE_DIR = Path("data/.cache")


def _pickle_cached(src_path: str, build_fn):
    """Return build_fn() result, cached to a pickle keyed on the source file's mtime."""
    src = Path(src_path)
    cache_file = _CACHE_DIR / (src.stem + ".pkl")
    if cache_file.exists() and src.exists() and cache_file.stat().st_mtime >= src.stat().st_mtime:
        with open(cache_file, "rb") as f:
            return pickle.load(f)
    obj = build_fn()
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with open(cache_file, "wb") as f:
        pickle.dump(obj, f)
    return obj

# ── Corpus registry: corpus_name -> local train/test file paths ───────────────
#
# corpus_name doubles as the domain ("wikitext"/"gutenberg"): same
# corpus_name is in-domain, a different one is out-of-domain.
CORPORA = {
    "wikitext":  {"train": "data/corpus_wikitext103.txt",
                  "test":  "data/test_corpus_wikitext103.txt"},
    "gutenberg": {"train": "data/corpus_gutenberg_1200books.txt",
                  "test":  "data/test_corpus_gutenberg.txt"},
}

# small/medium/large: take the first N sentences of the full training
# corpus (a fixed prefix, not a random sample, so a given tier is always
# the exact same sentences and needs no recorded seed).
CORPUS_SIZES = {"small": 0.2, "medium": 0.5, "large": 1.0}

# Gutenberg book count, chosen to land in the same ballpark as WikiText-103.
GUTENBERG_N_BOOKS = 1200


# ── Download + local cache (train corpora only; test files are small and
#    already checked into data/) ───────────────────────────────────────────────

def _download_wikitext(path: Path) -> None:
    print(f"[load_data] downloading WikiText-103 (train split) from HuggingFace...")
    from datasets import load_dataset
    ds = load_dataset("wikitext", "wikitext-103-raw-v1", split="train")

    def _keep(s: str) -> bool:
        s = s.strip()
        return bool(s) and not s.startswith("=")

    lines = [row["text"].strip() for row in ds if _keep(row["text"])]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[load_data] saved {path} ({len(lines)} lines)")


def _download_gutenberg(path: Path, n_books: int = GUTENBERG_N_BOOKS) -> None:
    print(f"[load_data] downloading {n_books} Gutenberg books from "
          f"HuggingFace (sedthh/gutenberg_english, streaming)...")
    from datasets import load_dataset
    ds = load_dataset("sedthh/gutenberg_english", split="train", streaming=True)

    lines: list[str] = []
    for i, book in enumerate(ds):
        if i >= n_books:
            break
        for line in book["TEXT"].split("\n"):
            line = line.strip()
            if line:
                lines.append(line)

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[load_data] saved {path} ({len(lines)} lines)")


_DOWNLOADERS = {
    "wikitext":  _download_wikitext,
    "gutenberg": _download_gutenberg,
}


def _ensure_train_file(corpus_name: str) -> str:
    """Download the full training corpus on first use; reuse the local file after."""
    path_str = CORPORA[corpus_name]["train"]
    path = Path(path_str)
    if not path.exists():
        _DOWNLOADERS[corpus_name](path)
    return path_str


def _load_local(path: str) -> list[str]:
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Corpus file not found: {path}\n"
            f"Make sure data/ has been populated (test files are committed; "
            f"train files are downloaded automatically on first use)."
        )
    return load_corpus(path)


def load_train_corpus(corpus_name: str, size: str = "large") -> list[str]:
    """
    Load corpus_name's training corpus (sentence list), sliced to `size`.
    Downloads and caches the full corpus on first use.

    Args:
        corpus_name: "wikitext" or "gutenberg"
        size: "small" (first 20%) / "medium" (first 50%) / "large" (100%)
    """
    if corpus_name not in CORPORA:
        raise ValueError(f"Unknown corpus_name: {corpus_name!r}, choose from {sorted(CORPORA)}")
    if size not in CORPUS_SIZES:
        raise ValueError(f"Unknown size: {size!r}, choose from {sorted(CORPUS_SIZES)}")

    full = _load_local(_ensure_train_file(corpus_name))
    frac = CORPUS_SIZES[size]
    if frac >= 1.0:
        return full
    return full[:round(len(full) * frac)]


def load_test_corpus(corpus_name: str) -> list[str]:
    """Load corpus_name's held-out test corpus (sentence list)."""
    if corpus_name not in CORPORA:
        raise ValueError(f"Unknown corpus_name: {corpus_name!r}, choose from {sorted(CORPORA)}")
    return _load_local(CORPORA[corpus_name]["test"])


# ═════════════════════════════════════════════════════════════════════════════
# MorphoLex
# ═════════════════════════════════════════════════════════════════════════════
#
# Source file is .xlsx, one sheet per PRS signature (e.g. "1-2-2" = 1
# prefix, 2 roots, 2 suffixes). The MorphoLexSegm field marks morphemes
# with nested brackets that carry role information but no boundary
# meaning themselves:
#   ( )   wraps a root's canonical form, e.g. (plic), (bio)
#   { }   wraps a "stem unit" (root + its immediate suffix) - just nesting
#   < >   wraps a prefix + the stem it governs - also just nesting
#
# So the morphemes are simply the letter runs inside these brackets, in
# left-to-right order, e.g.:
#   {(muli<(plic)>ate)}>ion>          -> [muli, plic, ate, ion]   (multiplications)
#   <im<{(contra)(vert)}>able>>y>     -> [im, contra, vert, able, y] (incontrovertibly)

def parse_morpholex_segmentation(raw: str) -> list[str]:
    """Extract the morpheme list from a MorphoLexSegm string (ignores all brackets)."""
    if not raw or not isinstance(raw, str):
        return []
    return [m.lower() for m in re.findall(r"[a-zA-Z]+", raw)]


def parse_morpholex_roles(raw: str) -> list[str]:
    """
    Role for each morpheme in a MorphoLexSegm string, in the same order
    as parse_morpholex_segmentation() (so the two outputs line up 1:1):
        root   - letter run wrapped in ( )
        prefix - letter run wrapped in < ... <
        suffix - letter run wrapped in > ... >
    """
    if not raw or not isinstance(raw, str):
        return []

    roles = []
    i, n = 0, len(raw)
    while i < n:
        if raw[i].isalpha():
            j = i
            while j < n and raw[j].isalpha():
                j += 1
            pre_char  = raw[i - 1] if i > 0 else ""
            post_char = raw[j]     if j < n else ""

            if pre_char == "(" and post_char == ")":
                roles.append("root")
            elif pre_char == "<" and post_char == "<":
                roles.append("prefix")
            elif post_char == ">" and post_char == ">":
                roles.append("suffix")
            i = j
        else:
            i += 1

    return roles


class MorphoLexDB:
    """
    Holds everything loaded from MorphoLex.

    Attributes:
        segmentation: dict[word -> list[morpheme]]
        roles:        dict[word -> list[role]], same length as segmentation,
                      role in {"prefix", "root", "suffix"}
        prefixes/suffixes/roots: sets of all known affixes/roots
    """

    def __init__(self):
        self.segmentation: dict[str, list[str]] = {}
        self.roles:        dict[str, list[str]] = {}
        self.prefixes:     set[str] = set()
        self.suffixes:     set[str] = set()
        self.roots:        set[str] = set()

    def get_morphemes(self, word: str) -> list[str]:
        return self.segmentation.get(word.lower(), [])

    def get_roles(self, word: str) -> list[str]:
        return self.roles.get(word.lower(), [])

    def get_prefixes_of(self, word: str) -> list[str]:
        m, r = self.get_morphemes(word), self.get_roles(word)
        return [m[i] for i in range(len(r)) if r[i] == "prefix"]

    def get_suffixes_of(self, word: str) -> list[str]:
        m, r = self.get_morphemes(word), self.get_roles(word)
        return [m[i] for i in range(len(r)) if r[i] == "suffix"]

    def get_roots_of(self, word: str) -> list[str]:
        m, r = self.get_morphemes(word), self.get_roles(word)
        return [m[i] for i in range(len(r)) if r[i] == "root"]

    def is_aligned(self, word: str) -> bool:
        """
        Whether this word's morphemes, concatenated, equal the word's
        actual spelling. MorphoLex morphemes are canonical forms (e.g.
        "implication" -> [in, plic, ate, ion], but the real spelling has
        "im" not "in" due to assimilation), so some words don't align —
        those should be filtered out before using them for any metric
        that compares against the word's real characters.
        """
        morphemes = self.get_morphemes(word)
        if not morphemes:
            return False
        return "".join(morphemes) == word.lower()

    def __repr__(self) -> str:
        import random
        sample_words = random.sample(list(self.segmentation.keys()),
                                     min(10, len(self.segmentation)))
        samples_str = "".join(f"\n    - {w}: {self.segmentation[w]} -> {self.roles[w]}"
                              for w in sample_words)
        return (
            f"<MorphoLexDB> loaded:\n"
            f"  - Total words: {len(self.segmentation)}\n"
            f"  - Sets: {len(self.prefixes)} prefixes, {len(self.suffixes)} suffixes, "
            f"{len(self.roots)} roots\n"
            f"  - Samples:{samples_str}"
        )


def filter_aligned_words(db: MorphoLexDB) -> set[str]:
    """Words whose morphemes concatenate back to the real spelling (see is_aligned)."""
    return {w for w in db.segmentation if db.is_aligned(w)}


def _clean_affix(raw: str) -> str:
    """"<mono<" or ">ing>" -> plain lowercase letters."""
    return re.sub(r"[^a-zA-Z]", "", raw).lower()


def _clean_root(raw: str) -> str:
    """"(play)" -> plain lowercase letters."""
    return re.sub(r"[^a-zA-Z]", "", raw).lower()


def load_morpholex(xlsx_path: str) -> MorphoLexDB:
    """Load MorphoLex into a MorphoLexDB (pickle-cached after the first parse)."""
    return _pickle_cached(xlsx_path, lambda: _parse_morpholex(xlsx_path))


def _parse_morpholex(xlsx_path: str) -> MorphoLexDB:
    """
    Parse the MorphoLex Excel file into a MorphoLexDB: every PRS-signature
    sheet (segmentation + roles) plus the "All prefixes"/"All
    suffixes"/"All roots" reference sheets.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Install openpyxl first: pip install openpyxl")

    db = MorphoLexDB()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_sheets = wb.sheetnames

    # Reference sheets: known prefix/suffix/root sets
    for sheet_name in all_sheets:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        sname_lower = sheet_name.strip().lower()

        if sname_lower in ("all prefixes", "all suffixes", "all roots"):
            morph_idx = next((i for i, h in enumerate(header) if "morpheme" in h), 0)
            clean = _clean_affix if sname_lower != "all roots" else _clean_root
            target = {"all prefixes": db.prefixes, "all suffixes": db.suffixes,
                      "all roots": db.roots}[sname_lower]
            for row in rows[1:]:
                if row[morph_idx]:
                    target.add(clean(str(row[morph_idx])))

    db.prefixes.discard("")
    db.suffixes.discard("")
    db.roots.discard("")

    # PRS-signature sheets: per-word segmentation + roles
    skip_sheets = {"all prefixes", "all suffixes", "all roots", "presentation"}
    word_count = 0

    for sheet_name in all_sheets:
        if sheet_name.strip().lower() in skip_sheets:
            continue

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c else "" for c in rows[0]]

        word_idx = next((i for i, h in enumerate(header) if h.lower() == "word"), None)
        segm_idx = next((i for i, h in enumerate(header) if "morpholexsegm" in h.lower()), None)
        if word_idx is None or segm_idx is None:
            continue

        for row in rows[1:]:
            if not row or len(row) <= max(word_idx, segm_idx):
                continue
            word, raw_segm = row[word_idx], row[segm_idx]
            if word is None or raw_segm is None:
                continue

            word     = str(word).strip().lower()
            raw_segm = str(raw_segm).strip()
            if not word or not raw_segm:
                continue

            morphemes = parse_morpholex_segmentation(raw_segm)
            roles     = parse_morpholex_roles(raw_segm)
            if not morphemes or len(morphemes) != len(roles):
                continue  # parse failed, skip

            db.segmentation[word] = morphemes
            db.roles[word]        = roles
            word_count += 1

    wb.close()
    return db


# ═════════════════════════════════════════════════════════════════════════════
# Morpho Challenge Gold Standard Segmentation
# ═════════════════════════════════════════════════════════════════════════════
#
# Morpho Challenge 2010 English gold standard (from Hutmegs/CELEX),
# surface-form (character-aligned) segmentation. Unlike MorphoLex, the
# morphemes here are literal substrings of the word's actual spelling.
#
# File format (tab-separated): word, then space-separated
# "surface:label" units; "," separates multiple candidate analyses
# (we keep the first); "~" marks "no surface form at this position".
#   accompanied\tac:ac_p compani:company_N ed:+PAST
#   aides-memoire\taides-memoire:aide-memoire_N ~:+PL

class GoldStdDB:
    """
    Holds the gold standard's surface segmentation.

    Attributes:
        segmentation: dict[word -> list[surface_morpheme]] (in order;
                      "~" placeholders are dropped)
        labels:       dict[word -> list[label]] (dictionary tags, unused
                      by current metrics but kept for reference)
    """

    def __init__(self):
        self.segmentation: dict[str, list[str]] = {}
        self.labels:       dict[str, list[str]] = {}

    def get_morphemes(self, word: str) -> list[str]:
        return self.segmentation.get(word.lower(), [])

    def get_labels(self, word: str) -> list[str]:
        return self.labels.get(word.lower(), [])

    def __repr__(self) -> str:
        import random
        all_words = list(self.segmentation.keys())
        sample_words = random.sample(all_words, min(10, len(all_words)))
        samples_str = "".join(f"\n    - {w}: {self.segmentation[w]} -> {self.get_labels(w)}"
                              for w in sample_words)
        return (
            f"<GoldStdDB> loaded:\n"
            f"  - Total words: {len(self.segmentation)}\n"
            f"  - Samples ({len(sample_words)}):{samples_str}"
        )


def parse_goldstd_line(raw_analysis: str) -> tuple[list[str], list[str]]:
    """
    Parse the morpheme-analysis part of one line (everything after the
    word + TAB), e.g. "ac:ac_p compani:company_N ed:+PAST" ->
    (["ac","compani","ed"], ["ac_p","company_N","+PAST"]).

    Rules: keep only the first candidate analysis (before any ","); skip
    units whose surface is "~" (no visible form); a literal hyphen
    surface ("-") is appended to the previous morpheme instead of kept
    as its own entry, so the concatenated surfaces still equal the
    original spelling (e.g. "armour-clad").
    """
    first_analysis = raw_analysis.split(",")[0].strip()
    surfaces: list[str] = []
    labels:   list[str] = []

    for unit in first_analysis.split():
        if ":" not in unit:
            continue
        surface, label = unit.split(":", 1)

        if surface == "~":
            continue
        if surface == "-":
            if surfaces:
                surfaces[-1] = surfaces[-1] + "-"
            continue

        surfaces.append(surface.lower())
        labels.append(label)

    return surfaces, labels


def load_goldstd(path: str) -> GoldStdDB:
    """Load a Morpho Challenge gold standard file (pickle-cached after first parse)."""
    return _pickle_cached(path, lambda: _parse_goldstd(path))


def _parse_goldstd(path: str) -> GoldStdDB:
    """Parse a Morpho Challenge gold standard segmentation file."""
    db = GoldStdDB()
    skipped = 0

    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line.strip():
                continue

            parts = line.split("\t")
            if len(parts) != 2:
                skipped += 1
                continue

            word, raw_analysis = parts
            word = word.strip().lower()
            surfaces, labels = parse_goldstd_line(raw_analysis)

            if not surfaces:
                skipped += 1
                continue

            db.segmentation[word] = surfaces
            db.labels[word]       = labels

    return db


# ═════════════════════════════════════════════════════════════════════════════
# Self-checks
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # (raw string, word, expected morphemes, expected roles)
    test_cases = [
        ("{(allegor)>ic>}al>>ly>", "allegorically",
         ["allegor","ic","al","ly"], ["root","suffix","suffix","suffix"]),
        ("{(whim)>y>}al>>ity>", "whimsicality",
         ["whim","y","al","ity"], ["root","suffix","suffix","suffix"]),
        ("{(transit)}ion>>al>>ly>", "transitionally",
         ["transit","ion","al","ly"], ["root","suffix","suffix","suffix"]),
        ("{(weigh)>t>}y>>ness>", "weightiness",
         ["weigh","t","y","ness"], ["root","suffix","suffix","suffix"]),
        ("<im<{(contra)(vert)}>able>>y>", "incontrovertibly",
         ["im","contra","vert","able","y"], ["prefix","root","root","suffix","suffix"]),
        ("<un<{(trust)}{(worth)}>y>>ness>", "untrustworthiness",
         ["un","trust","worth","y","ness"], ["prefix","root","root","suffix","suffix"]),
        ("<un<{(equi)(voc)>al>}>ly>", "unequivocally",
         ["un","equi","voc","al","ly"], ["prefix","root","root","suffix","suffix"]),
        ("<micro<(cyto){(chem)>ist>}>ory>", "microcytochemistry",
         ["micro","cyto","chem","ist","ory"], ["prefix","root","root","suffix","suffix"]),
        ("<auto<{(bio)(graph)>ic>}>al>", "autobiographical",
         ["auto","bio","graph","ic","al"], ["prefix","root","root","suffix","suffix"]),
        ("<de<{(myth)}(log)>iz>>ion>", "demythologization",
         ["de","myth","log","iz","ion"], ["prefix","root","root","suffix","suffix"]),
        ("{(mal)}{<a<(minister)>ate>}>ion>", "maladministration",
         ["mal","a","minister","ate","ion"], ["root","prefix","root","suffix","suffix"]),
        ("{(kin)}{<es<(thet)>ic>}>ly>", "kinesthetically",
         ["kin","es","thet","ic","ly"], ["root","prefix","root","suffix","suffix"]),
        ("<non<{(myth)(log)>ic>}>al>", "nonmythological",
         ["non","myth","log","ic","al"], ["prefix","root","root","suffix","suffix"]),
    ]

    print("=" * 70)
    print("  Testing parse_morpholex_segmentation + parse_morpholex_roles")
    print("=" * 70)

    passed = failed = 0
    for raw, word, exp_morphemes, exp_roles in test_cases:
        got_morphemes = parse_morpholex_segmentation(raw)
        got_roles     = parse_morpholex_roles(raw)
        ok = got_morphemes == exp_morphemes and got_roles == exp_roles
        print(f"\n  {'OK' if ok else 'FAIL'}  {word}")
        print(f"       raw      : {raw}")
        print(f"       morphemes: {got_morphemes}" + ("" if ok else f"  (expected {exp_morphemes})"))
        print(f"       roles    : {got_roles}" + ("" if ok else f"  (expected {exp_roles})"))
        passed += ok
        failed += (not ok)

    print(f"\n{'=' * 70}")
    print(f"  {passed} passed, {failed} failed.")
    print("=" * 70)

    # GoldStd: parse_goldstd_line
    goldstd_test_cases = [
        ("ac:ac_p compani:company_N ed:+PAST", ["ac", "compani", "ed"]),
        ("advers:adverse_A ari:ary_s es:+PL", ["advers", "ari", "es"]),
        ("a:a_p fire:fire_N, a:a_p fire:fire_N", ["a", "fire"]),            # multiple candidates
        ("aides-memoire:aide-memoire_N ~:+PL", ["aides-memoire"]),         # ~ skipped
        ("anarch:anarchy_N ~:ism_s ist:ist_s s:+PL", ["anarch", "ist", "s"]),
        ("arm:arm_V our:our_s -:~ clad:clad_A", ["arm", "our-", "clad"]),  # hyphen reattached
        ("accent:accent_N s:+PL ':+GEN", ["accent", "s", "'"]),
    ]

    print(f"\n{'=' * 64}")
    print("  Testing parse_goldstd_line")
    print("=" * 64)
    gs_passed = gs_failed = 0
    for raw, expected in goldstd_test_cases:
        surfaces, labels = parse_goldstd_line(raw)
        ok = surfaces == expected
        print(f"\n  {'OK' if ok else 'FAIL'}  raw: {raw}")
        print(f"     surfaces: {surfaces}" + ("" if ok else f"  (expected {expected})"))
        gs_passed += ok
        gs_failed += (not ok)

    print(f"\n{'='*64}")
    print(f"  {gs_passed} passed, {gs_failed} failed.")

    # Optional: character-alignment check against a real goldstd file
    import sys
    if len(sys.argv) > 1:
        print(f"\n{'='*64}")
        print(f"  GoldStd alignment check: {sys.argv[1]}")
        print(f"{'='*64}")
        gs_db = load_goldstd(sys.argv[1])
        aligned = sum(1 for w, m in gs_db.segmentation.items() if "".join(m) == w)
        total = len(gs_db.segmentation)
        print(f"  fully aligned: {aligned}/{total} ({aligned/total*100:.1f}%)")
