"""
run_eval.py
===========
Interactive evaluation entry point.

Three options, each prompting for the parameters it needs up front (no
back-and-forth menus):

  1. Evaluate one tokenizer  - full basic-statistics + morphological
     report for a single (algo, corpus, size, vocab_size[, min_frequency])
     config. Trains it if not already cached.
  2. Compare BPE vs WordPiece vs gold standard - same (corpus, size,
     vocab_size, min_frequency) used for both algorithms; prints a
     side-by-side metric table plus a few sample-word segmentations.
  3. Case study - interactive prompt: type a word/sentence, see how
     BPE / WordPiece / the gold standards each segment it.

Tokenization results are cached to disk so re-running the same
(tokenizer config, text) doesn't re-tokenize:
  results/reference_free/{config_id}__{corpus}.json   - test-corpus runs
  results/reference_based/{config_id}.json             - gold-standard word-list runs

Usage:
    python run_eval.py
"""

from __future__ import annotations
import json
import random
from datetime import datetime
from pathlib import Path

import openpyxl

from load_data import (
    CORPORA, CORPUS_SIZES, load_train_corpus, load_test_corpus,
    load_morpholex, filter_aligned_words, load_goldstd,
)
from metrics import (
    clean_token, basic_stats, extract_words, TokenizationCache, CachedTokenizer,
    run_morpholex_metrics, run_goldstd_metrics, gold_avg_tokens_per_wordtype,
)
from train_tokenizer import ALGO_CLASSES, get_or_train, vocab_path

RESULTS_DIR          = Path("results")
REFERENCE_FREE_DIR   = RESULTS_DIR / "reference_free"
REFERENCE_BASED_DIR  = RESULTS_DIR / "reference_based"
METRICS_XLSX         = RESULTS_DIR / "metrics_summary.xlsx"

MORPHOLEX_PATH = "data/MorphoLEX_en.xlsx"
GOLDSTD_PATH   = "data/goldstd_combined.segmentation.eng"

N_SEEN, N_UNSEEN = 1000, 500
N_SAMPLE_WORDS   = 8

# Every metric produced by options 1 and 2 is appended as one row to a single
# workbook (results/metrics_summary.xlsx). Rows with the same KEY_FIELDS are
# overwritten in place, so re-running a config updates rather than duplicates.
METRICS_KEY_FIELDS = ["algorithm", "corpus", "size", "vocab_size",
                      "min_frequency", "row_type", "tested_on", "split"]
METRICS_FIELDS = METRICS_KEY_FIELDS + [
    "timestamp", "actual_vocab_size", "in_domain", "tokenize_sec",
    # reference-free (basic statistics)
    "fertility", "avg_tokens_wordtype", "avg_tokens_wordtype_high",
    "avg_tokens_wordtype_low", "avg_token_length", "unk_rate",
    "intact_rate", "intact_rate_high", "intact_rate_low",
    # reference-based (morphological)
    "gold_avg_tokens_wordtype", "boundary_P", "boundary_R", "boundary_F1",
    "boundary_macro_F1", "morpheme_recovery", "over_seg_rate", "under_seg_rate",
    "prefix_recall", "suffix_recall", "consistency",
]


# ─────────────────────────────────────────
# Gold standard DBs (loaded once, reused across menu options)
# ─────────────────────────────────────────

_dbs_cache: dict = {}


def get_dbs():
    """Load MorphoLex + GoldStd once per process (cached for the session)."""
    if "morpholex" not in _dbs_cache:
        _dbs_cache["morpholex"] = load_morpholex(MORPHOLEX_PATH)
    if "goldstd" not in _dbs_cache:
        _dbs_cache["goldstd"] = load_goldstd(GOLDSTD_PATH)
    return _dbs_cache["morpholex"], _dbs_cache["goldstd"]


def train_vocab_of(corpus: list[str]) -> set[str]:
    """Set of word forms appearing in the training corpus (for seen/unseen splits)."""
    vocab: set[str] = set()
    for s in corpus:
        vocab.update(s.lower().split())
    return vocab


def build_morpholex_test_words(morpholex_db, train_vocab: set[str],
                               n_seen: int = N_SEEN,
                               n_unseen: int = N_UNSEEN) -> tuple[list[str], list[str]]:
    """MorphoLex test words: only spelling-aligned words, split by train-vocab membership."""
    aligned_words = filter_aligned_words(morpholex_db)
    candidate_words = {w for w in aligned_words if len(morpholex_db.segmentation[w]) >= 2}

    # Sort before sampling: set iteration order varies by process (hash seed),
    # so without this the chosen words (and thus the cached tokenizations keyed
    # by config) would differ run to run. Sorting makes selection reproducible.
    seen   = sorted(w for w in candidate_words if w in train_vocab)
    unseen = sorted(w for w in candidate_words if w not in train_vocab)

    random.seed(42)
    seen   = random.sample(seen,   min(n_seen,   len(seen)))
    unseen = random.sample(unseen, min(n_unseen, len(unseen)))
    return seen, unseen


def build_goldstd_test_words(goldstd_db, train_vocab: set[str],
                             n_seen: int = N_SEEN,
                             n_unseen: int = N_UNSEEN) -> tuple[list[str], list[str]]:
    """GoldStd test words: 100% surface-aligned, no filtering needed."""
    candidate_words = {w for w, m in goldstd_db.segmentation.items() if len(m) >= 2}

    # Sort before sampling (see build_morpholex_test_words) for reproducibility.
    seen   = sorted(w for w in candidate_words if w in train_vocab)
    unseen = sorted(w for w in candidate_words if w not in train_vocab)

    random.seed(42)
    seen   = random.sample(seen,   min(n_seen,   len(seen)))
    unseen = random.sample(unseen, min(n_unseen, len(unseen)))
    return seen, unseen


# ─────────────────────────────────────────
# Tokenization-result caching (reference_free = test corpus,
# reference_based = gold-standard word lists)
# ─────────────────────────────────────────

def _load_or_build_cache(path: Path, tok, words: list[str]) -> TokenizationCache:
    """Load the cached word->tokens result if present, else tokenize and save it."""
    if path.exists():
        with open(path, encoding="utf-8") as f:
            return TokenizationCache.load_from_dict(json.load(f))
    cache = TokenizationCache.build(tok, words)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache.save_to_dict(), f, ensure_ascii=False)
    return cache


def reference_free_cache(tok, config_id: str, corpus_name: str,
                         corpus: list[str]) -> TokenizationCache:
    """
    Cache the tokenization of a test corpus (reused across runs). Walks
    every word occurrence (repeats included) so total_sec reflects
    processing the whole corpus, not just its unique vocabulary.
    """
    path = REFERENCE_FREE_DIR / f"{config_id}__{corpus_name}.json"
    return _load_or_build_cache(path, tok, extract_words(corpus))


def reference_based_cache(tok, config_id: str, words: list[str]) -> TokenizationCache:
    """Cache the tokenization of a gold-standard word list (reused across runs)."""
    path = REFERENCE_BASED_DIR / f"{config_id}.json"
    return _load_or_build_cache(path, tok, words)


def _config_id(algo: str, corpus_name: str, size: str, vocab_size: int,
               min_frequency: int | None) -> str:
    return vocab_path(algo, corpus_name, size, vocab_size,
                      min_frequency if algo == "wp" else None).stem


# ─────────────────────────────────────────
# Console report helpers
# ─────────────────────────────────────────

def _hr(char: str = "-", width: int = 60) -> None:
    print(char * width)


def _basic_stats_row_values(res: dict) -> dict:
    tld = res["token_length_dist"]
    total_chars = sum(l * c for l, c in tld.items())
    total_count = sum(tld.values())
    return {
        "fertility":              res["fertility"]["fertility"],
        "avg_tokens_wordtype":     res["wordtype_stats"]["all"]["avg_tokens"],
        "avg_tokens_wordtype_high": res["wordtype_stats"]["high_freq"]["avg_tokens"],
        "avg_tokens_wordtype_low": res["wordtype_stats"]["low_freq"]["avg_tokens"],
        "avg_token_length":        round(total_chars / total_count, 4) if total_count else 0.0,
        "unk_rate":                res["unk_rate"],
        "intact_rate":             round(res["intact_token_stats"]["all"]["intact_pct"] / 100, 4),
        "intact_rate_high":        round(res["intact_token_stats"]["high_freq"]["intact_pct"] / 100, 4),
        "intact_rate_low":         round(res["intact_token_stats"]["low_freq"]["intact_pct"] / 100, 4),
    }


# ─────────────────────────────────────────
# Single Excel metrics summary (results/metrics_summary.xlsx)
# ─────────────────────────────────────────

def _config_meta(algo_label: str, corpus_name: str, size: str, vocab_size: int,
                 min_frequency, actual_vocab_size: int) -> dict:
    """The shared identifying columns for a row (filled into both row types)."""
    return {
        "algorithm": algo_label, "corpus": corpus_name, "size": size,
        "vocab_size": vocab_size,
        "min_frequency": min_frequency if min_frequency is not None else "",
        "actual_vocab_size": actual_vocab_size,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def append_metrics_rows(rows: list[dict]) -> None:
    """Append/overwrite metric rows in results/metrics_summary.xlsx (dedup by key)."""
    if not rows:
        return
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    existing: list[dict] = []
    if METRICS_XLSX.exists():
        ws = openpyxl.load_workbook(METRICS_XLSX).active
        data = list(ws.iter_rows(values_only=True))
        if data:
            header = list(data[0])
            existing = [dict(zip(header, r)) for r in data[1:]]

    def key(r: dict) -> tuple:
        # Empty cells read back from xlsx come in as None; normalize so a
        # re-run's "" key still matches the stored row and overwrites it.
        return tuple("" if r.get(k) is None else str(r.get(k)) for k in METRICS_KEY_FIELDS)

    new_keys = {key(r) for r in rows}
    merged = [r for r in existing if key(r) not in new_keys] + rows

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "metrics"
    ws.append(METRICS_FIELDS)
    for r in merged:
        ws.append([r.get(f, "") for f in METRICS_FIELDS])
    wb.save(METRICS_XLSX)
    print(f"\n  Metrics written to {METRICS_XLSX}")


# ─────────────────────────────────────────
# Option 1: evaluate one tokenizer
# ─────────────────────────────────────────

def evaluate_one(algo: str, corpus_name: str, size: str, vocab_size: int,
                 min_frequency: int = 1) -> None:
    tok = get_or_train(algo, corpus_name, size, vocab_size, min_frequency)
    config_id = _config_id(algo, corpus_name, size, vocab_size, min_frequency)
    algo_label = "BPE" if algo == "bpe" else "WordPiece"
    home_domain = corpus_name

    print()
    _hr("=")
    print(f"  {algo_label}  |  corpus={corpus_name}  size={size}  vocab_size={vocab_size}"
          + (f"  min_frequency={min_frequency}" if algo == "wp" else ""))
    _hr("=")
    print(f"  actual vocab size: {tok.vocab_size()}    merges: {len(tok.merges)}")

    # ── Basic statistics, in-domain and out-of-domain ──────────────────────
    print("\n  Basic statistics (test corpora)")
    _hr()
    domains = sorted(CORPORA, key=lambda d: d != home_domain)  # home domain first
    rows = {}
    res_by_domain = {}
    for domain_name in domains:
        corpus = load_test_corpus(domain_name)
        cache = reference_free_cache(tok, config_id, domain_name, corpus)
        res = basic_stats(tok, corpus, cache=cache)
        res_by_domain[domain_name] = res
        rows[domain_name] = _basic_stats_row_values(res)

    tag = {d: ("in-domain" if d == home_domain else "out-of-domain") for d in domains}
    header_cols = [f"{d} ({tag[d]})" for d in domains]
    print(f"  {'test corpus:':<28}" + "".join(f"{c:>22}" for c in header_cols))
    for key in ["fertility", "avg_tokens_wordtype", "avg_tokens_wordtype_high",
               "avg_tokens_wordtype_low", "avg_token_length", "unk_rate",
               "intact_rate", "intact_rate_high", "intact_rate_low"]:
        print(f"  {key:<28}" + "".join(f"{rows[d][key]:>22}" for d in domains))

    # ── Morphological metrics, seen/unseen splits ───────────────────────────
    morpholex_db, goldstd_db = get_dbs()
    train_vocab = train_vocab_of(load_train_corpus(corpus_name, size))
    morpholex_seen, morpholex_unseen = build_morpholex_test_words(morpholex_db, train_vocab)
    goldstd_seen, goldstd_unseen     = build_goldstd_test_words(goldstd_db, train_vocab)

    combined_words = sorted(set(morpholex_seen) | set(morpholex_unseen) |
                            set(goldstd_seen) | set(goldstd_unseen))
    ref_cache = reference_based_cache(tok, config_id, combined_words)
    cached_tok = CachedTokenizer(ref_cache, tok.vocab_size())

    print("\n  Morphological metrics (gold standard: GoldStd boundary metrics, "
          "MorphoLex affix/consistency)")
    _hr()
    print(f"  {'':<28}{'seen':>14}{'unseen':>14}")
    split_vals = {}
    for split_name, gs_words, ml_words in [
        ("seen", goldstd_seen, morpholex_seen),
        ("unseen", goldstd_unseen, morpholex_unseen),
    ]:
        gres = run_goldstd_metrics(cached_tok, algo_label, gs_words, goldstd_db)
        mres = run_morpholex_metrics(cached_tok, ml_words, morpholex_db)
        bprf = gres["boundary_prf"]
        ou   = gres["over_under"]
        ad   = mres["affix_detection"]
        split_vals[split_name] = {
            "boundary_macro_F1": bprf.macro_F1,
            "boundary_P": bprf.micro_P, "boundary_R": bprf.micro_R, "boundary_F1": bprf.micro_F1,
            "morpheme_recovery": gres["morpheme_recovery"],
            "over_seg_rate": ou["over_seg_rate"], "under_seg_rate": ou["under_seg_rate"],
            "prefix_recall": ad["prefix_recall"], "suffix_recall": ad["suffix_recall"],
            "consistency": mres["consistency"]["mean_consistency"],
            "gold_avg_tokens_wordtype": gold_avg_tokens_per_wordtype(goldstd_db, gs_words),
        }
    seen_vals, unseen_vals = split_vals["seen"], split_vals["unseen"]

    for key, label in [
        ("boundary_P", "Boundary Precision (micro)"),
        ("boundary_R", "Boundary Recall (micro)"),
        ("boundary_F1", "Boundary F1 (micro)"),
        ("boundary_macro_F1", "Boundary F1 (macro)"),
        ("morpheme_recovery", "Morpheme Recovery Rate"),
        ("over_seg_rate", "Over-segmentation rate"),
        ("under_seg_rate", "Under-segmentation rate"),
        ("prefix_recall", "Prefix Recall"),
        ("suffix_recall", "Suffix Recall"),
        ("consistency", "Root Consistency"),
        ("gold_avg_tokens_wordtype", "Gold avg tokens/wordtype"),
    ]:
        print(f"  {label:<28}{round(seen_vals[key],4):>14}{round(unseen_vals[key],4):>14}")

    # ── Write every metric to the single Excel summary ──────────────────────
    meta = _config_meta(algo_label, corpus_name, size, vocab_size,
                        min_frequency if algo == "wp" else None, tok.vocab_size())
    excel_rows = []
    for domain_name in domains:
        excel_rows.append({**meta, "row_type": "basic", "tested_on": domain_name,
                           "split": "", "in_domain": domain_name == home_domain,
                           "tokenize_sec": res_by_domain[domain_name]["total_sec"],
                           **rows[domain_name]})
    for split_name in ("seen", "unseen"):
        excel_rows.append({**meta, "row_type": "morph", "tested_on": "",
                           "split": split_name, **split_vals[split_name]})
    append_metrics_rows(excel_rows)
    print()


# ─────────────────────────────────────────
# Option 2: compare BPE vs WordPiece vs gold standard
# ─────────────────────────────────────────

def compare_with_gold(corpus_name: str, size: str, vocab_size: int, min_frequency: int) -> None:
    bpe_tok = get_or_train("bpe", corpus_name, size, vocab_size)
    wp_tok  = get_or_train("wp",  corpus_name, size, vocab_size, min_frequency)
    bpe_id  = _config_id("bpe", corpus_name, size, vocab_size, None)
    wp_id   = _config_id("wp",  corpus_name, size, vocab_size, min_frequency)

    # ── Basic statistics, in-domain only ────────────────────────────────────
    test_corpus = load_test_corpus(corpus_name)
    bpe_basic = basic_stats(bpe_tok, test_corpus,
                            cache=reference_free_cache(bpe_tok, bpe_id, corpus_name, test_corpus))
    wp_basic  = basic_stats(wp_tok, test_corpus,
                            cache=reference_free_cache(wp_tok, wp_id, corpus_name, test_corpus))
    bpe_row = _basic_stats_row_values(bpe_basic)
    wp_row  = _basic_stats_row_values(wp_basic)

    # ── Morphological metrics, seen+unseen combined ────────────────────────
    morpholex_db, goldstd_db = get_dbs()
    train_vocab = train_vocab_of(load_train_corpus(corpus_name, size))
    morpholex_seen, morpholex_unseen = build_morpholex_test_words(morpholex_db, train_vocab)
    goldstd_seen, goldstd_unseen     = build_goldstd_test_words(goldstd_db, train_vocab)
    morpholex_words = sorted(set(morpholex_seen) | set(morpholex_unseen))
    goldstd_words   = sorted(set(goldstd_seen) | set(goldstd_unseen))
    combined_words  = sorted(set(morpholex_words) | set(goldstd_words))

    bpe_cached = CachedTokenizer(reference_based_cache(bpe_tok, bpe_id, combined_words),
                                 bpe_tok.vocab_size())
    wp_cached  = CachedTokenizer(reference_based_cache(wp_tok, wp_id, combined_words),
                                 wp_tok.vocab_size())

    bpe_gold = run_goldstd_metrics(bpe_cached, "BPE", goldstd_words, goldstd_db)
    wp_gold  = run_goldstd_metrics(wp_cached, "WordPiece", goldstd_words, goldstd_db)
    bpe_morph = run_morpholex_metrics(bpe_cached, morpholex_words, morpholex_db)
    wp_morph  = run_morpholex_metrics(wp_cached, morpholex_words, morpholex_db)
    gold_avg_tokens = gold_avg_tokens_per_wordtype(goldstd_db, goldstd_words)

    # ── Print comparison table ──────────────────────────────────────────────
    print()
    _hr("=")
    print(f"  BPE vs WordPiece  |  corpus={corpus_name}  size={size}  "
          f"vocab_size={vocab_size}  min_frequency={min_frequency}  (in-domain test set)")
    _hr("=")
    print(f"  {'':<28}{'BPE':>14}{'WordPiece':>14}{'Gold':>14}")
    for key, label in [
        ("fertility", "Fertility"),
        ("avg_tokens_wordtype", "Avg tokens / wordtype"),
        ("avg_tokens_wordtype_high", "  (high-frequency)"),
        ("avg_tokens_wordtype_low", "  (low-frequency)"),
        ("avg_token_length", "Avg token length"),
        ("unk_rate", "UNK rate"),
        ("intact_rate", "Intact-token rate"),
        ("intact_rate_high", "  (high-frequency)"),
        ("intact_rate_low", "  (low-frequency)"),
    ]:
        gold_col = round(gold_avg_tokens, 4) if key == "avg_tokens_wordtype" else ""
        print(f"  {label:<28}{bpe_row[key]:>14}{wp_row[key]:>14}{gold_col!s:>14}")

    print()
    for key, label, src_bpe, src_wp in [
        ("micro_P", "Boundary Precision (micro)", bpe_gold["boundary_prf"], wp_gold["boundary_prf"]),
        ("micro_R", "Boundary Recall (micro)", bpe_gold["boundary_prf"], wp_gold["boundary_prf"]),
        ("micro_F1", "Boundary F1 (micro)", bpe_gold["boundary_prf"], wp_gold["boundary_prf"]),
    ]:
        print(f"  {label:<28}{round(getattr(src_bpe, key), 4):>14}{round(getattr(src_wp, key), 4):>14}")
    print(f"  {'Morpheme Recovery Rate':<28}{bpe_gold['morpheme_recovery']:>14}{wp_gold['morpheme_recovery']:>14}")
    print(f"  {'Over-segmentation rate':<28}{bpe_gold['over_under']['over_seg_rate']:>14}"
          f"{wp_gold['over_under']['over_seg_rate']:>14}")
    print(f"  {'Under-segmentation rate':<28}{bpe_gold['over_under']['under_seg_rate']:>14}"
          f"{wp_gold['over_under']['under_seg_rate']:>14}")
    print(f"  {'Prefix Recall':<28}{bpe_morph['affix_detection']['prefix_recall']:>14}"
          f"{wp_morph['affix_detection']['prefix_recall']:>14}")
    print(f"  {'Suffix Recall':<28}{bpe_morph['affix_detection']['suffix_recall']:>14}"
          f"{wp_morph['affix_detection']['suffix_recall']:>14}")
    print(f"  {'Root Consistency':<28}{bpe_morph['consistency']['mean_consistency']:>14}"
          f"{wp_morph['consistency']['mean_consistency']:>14}")
    print(f"  {'Gold avg tokens/wordtype':<28}{'':>14}{'':>14}{round(gold_avg_tokens,4):>14}")

    # ── Write both tokenizers' metrics to the single Excel summary ──────────
    excel_rows = []
    for algo_label, mf, basic, brow, gold, morph, vsz in [
        ("BPE", None, bpe_basic, bpe_row, bpe_gold, bpe_morph, bpe_tok.vocab_size()),
        ("WordPiece", min_frequency, wp_basic, wp_row, wp_gold, wp_morph, wp_tok.vocab_size()),
    ]:
        meta = _config_meta(algo_label, corpus_name, size, vocab_size, mf, vsz)
        bprf = gold["boundary_prf"]
        excel_rows.append({**meta, "row_type": "basic", "tested_on": corpus_name,
                           "split": "", "in_domain": True,
                           "tokenize_sec": basic["total_sec"], **brow})
        excel_rows.append({**meta, "row_type": "morph", "tested_on": "", "split": "all",
                           "boundary_P": bprf.micro_P, "boundary_R": bprf.micro_R,
                           "boundary_F1": bprf.micro_F1, "boundary_macro_F1": bprf.macro_F1,
                           "morpheme_recovery": gold["morpheme_recovery"],
                           "over_seg_rate": gold["over_under"]["over_seg_rate"],
                           "under_seg_rate": gold["over_under"]["under_seg_rate"],
                           "prefix_recall": morph["affix_detection"]["prefix_recall"],
                           "suffix_recall": morph["affix_detection"]["suffix_recall"],
                           "consistency": morph["consistency"]["mean_consistency"],
                           "gold_avg_tokens_wordtype": gold_avg_tokens})
    append_metrics_rows(excel_rows)

    # ── Sample words: 3-way segmentation comparison ─────────────────────────
    print(f"\n  Sample segmentations (BPE vs WordPiece vs Gold):")
    _hr()
    sample = random.sample(goldstd_words, min(N_SAMPLE_WORDS, len(goldstd_words)))
    for w in sample:
        bpe_clean = "|".join(clean_token(t) for t in bpe_tok.tokenize(w))
        wp_clean  = "|".join(clean_token(t) for t in wp_tok.tokenize(w))
        gold_str  = "|".join(goldstd_db.segmentation.get(w, []))
        print(f"  {w:<22} BPE={bpe_clean:<28} WordPiece={wp_clean:<28} Gold={gold_str}")
    print()


# ─────────────────────────────────────────
# Option 3: case study (interactive)
# ─────────────────────────────────────────

def case_study(corpus_name: str, size: str, vocab_size: int, min_frequency: int) -> None:
    bpe_tok = get_or_train("bpe", corpus_name, size, vocab_size)
    wp_tok  = get_or_train("wp",  corpus_name, size, vocab_size, min_frequency)
    morpholex_db, goldstd_db = get_dbs()

    print("\nCase study: type a word or sentence to compare BPE / WordPiece / gold "
          "segmentations. Type 'quit' to exit.")

    while True:
        try:
            text = input("\n>>> ").strip()
        except (KeyboardInterrupt, EOFError):
            break
        if not text or text.lower() == "quit":
            break

        if " " in text:
            print(f"  {'BPE':<12}: {bpe_tok.tokenize(text)}")
            print(f"  {'WordPiece':<12}: {wp_tok.tokenize(text)}")
            continue

        word = text.lower()
        bpe_tokens = bpe_tok.tokenize(word)
        wp_tokens  = wp_tok.tokenize(word)
        print(f"  {'BPE':<12}: {bpe_tokens}  ->  {'|'.join(clean_token(t) for t in bpe_tokens)}")
        print(f"  {'WordPiece':<12}: {wp_tokens}  ->  {'|'.join(clean_token(t) for t in wp_tokens)}")

        if word in goldstd_db.segmentation:
            print(f"  {'GoldStd':<12}: {'|'.join(goldstd_db.segmentation[word])}")
        else:
            print(f"  {'GoldStd':<12}: (no entry)")

        if word in morpholex_db.segmentation:
            labeled = [f"{m}({r})" for m, r in zip(morpholex_db.segmentation[word],
                                                    morpholex_db.roles.get(word, []))]
            print(f"  {'MorphoLex':<12}: {labeled}")
        else:
            print(f"  {'MorphoLex':<12}: (no entry)")


# ─────────────────────────────────────────
# Parameter prompts
# ─────────────────────────────────────────

def _prompt(msg: str, choices: set | None = None, cast=str, default=None):
    suffix = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"{msg}{suffix}: ").strip()
        if not raw:
            if default is not None:
                return default
            continue
        try:
            value = cast(raw)
        except ValueError:
            print("  Invalid input, try again.")
            continue
        if choices and value not in choices:
            print(f"  Choose one of: {sorted(choices)}")
            continue
        return value


def _prompt_corpus_size_vocab(need_min_frequency_for: set[str]) -> tuple[str, str, int, dict[str, int]]:
    """
    Prompt for corpus/size/vocab_size, then a min_frequency per algo in
    need_min_frequency_for (e.g. {"wp"} for option 1 with algo=wp, or
    {"wp"} always for options 2/3 which evaluate both algorithms but
    only wp needs the parameter).
    """
    print(f"  corpus options: {sorted(CORPORA)}")
    corpus_name = _prompt("  corpus", choices=set(CORPORA), cast=str.lower)
    print(f"  size options: {sorted(CORPUS_SIZES)}")
    size = _prompt("  size", choices=set(CORPUS_SIZES), cast=str.lower)
    vocab_size = _prompt("  vocab_size", cast=int)

    min_frequency = {}
    if need_min_frequency_for:
        min_frequency["wp"] = _prompt("  min_frequency (WordPiece only)", cast=int, default=1)
    return corpus_name, size, vocab_size, min_frequency


# ─────────────────────────────────────────
# Main menu
# ─────────────────────────────────────────

def main() -> None:
    print("=" * 60)
    print("  Tokenizer evaluation")
    print("=" * 60)

    while True:
        print("\nChoose an action:")
        print("  1. Evaluate one tokenizer (trains it if not already cached)")
        print("  2. Compare BPE vs WordPiece vs gold standard")
        print("  3. Case study (interactive word/sentence comparison)")
        print("  4. Exit")
        choice = input("> ").strip()

        if choice == "1":
            print(f"\nParameters needed: algo, corpus, size, vocab_size, "
                  f"min_frequency (WordPiece only)")
            print(f"  algo options: {sorted(ALGO_CLASSES)}")
            algo = _prompt("  algo", choices=set(ALGO_CLASSES), cast=str.lower)
            corpus_name, size, vocab_size, mf = _prompt_corpus_size_vocab({"wp"} if algo == "wp" else set())
            evaluate_one(algo, corpus_name, size, vocab_size, mf.get("wp", 1))

        elif choice == "2":
            print(f"\nParameters needed: corpus, size, vocab_size, min_frequency "
                  f"(used by WordPiece only; BPE ignores it)")
            corpus_name, size, vocab_size, mf = _prompt_corpus_size_vocab({"wp"})
            compare_with_gold(corpus_name, size, vocab_size, mf["wp"])

        elif choice == "3":
            print(f"\nParameters needed: corpus, size, vocab_size, min_frequency "
                  f"(used by WordPiece only; BPE ignores it)")
            corpus_name, size, vocab_size, mf = _prompt_corpus_size_vocab({"wp"})
            case_study(corpus_name, size, vocab_size, mf["wp"])

        elif choice == "4":
            break
        else:
            print("Choose 1-4.")


if __name__ == "__main__":
    import sys
    for _stream in (sys.stdout, sys.stderr):
        if hasattr(_stream, "reconfigure"):
            _stream.reconfigure(encoding="utf-8", errors="replace")
    try:
        main()
    except (KeyboardInterrupt, EOFError):
        print("\nExiting.")
